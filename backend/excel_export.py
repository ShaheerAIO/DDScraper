"""
Excel export: loads the template workbook, clears data rows,
and fills all 14 sheets from the mapped model.
"""

import openpyxl


# Column header -> key mapping per sheet (order matters - matches template columns)
SHEET_COLUMNS = {
    "Menu": ["id", "menuName", "posDisplayName", "posButtonColor", "picture", "sortOrder"],
    "Category": [
        "id", "categoryName", "posDisplayName", "kdsDisplayName", "color",
        "image", "kioskImage", "parentCategoryId", "tagIds", "menuIds", "sortOrder",
    ],
    "Item": [
        "id", "itemName", "posDisplayName", "kdsName", "itemDescription",
        "itemPicture", "onlineImage", "landscapeImage", "thirdPartyImage",
        "kioskItemImage", "itemPrice", "taxLinkedWithParentSetting",
        "calculatePricesWithTaxIncluded", "takeoutException", "stockStatus",
        "stockValue", "orderQuantityLimit", "minLimit", "maxLimit", "noMaxLimit",
        "stationIds", "preparationTime", "calories", "tagIds",
        "inheritTagsFromCategory", "saleCategory", "allergenIds",
        "inheritModifiersFromCategory", "addonIds", "isSpecialRequest",
    ],
    "Item Modifiers": ["itemId", "modifierId", "sortOrder"],
    "Category ModifierGroups": ["categoryId", "modifierGroupId", "sortOrder"],
    "Category Modifiers": ["categoryId", "modifierId", "sortOrder"],
    "Category Items": ["id", "categoryId", "itemId", "sortOrder"],
    "Item Modifier Group": ["itemId", "modifierGroupId", "sortOrder"],
    "Modifier Group": ["id", "groupName", "posDisplayName", "onPrem", "offPrem", "modifierIds"],
    "Modifier": [
        "id", "modifierName", "posDisplayName", "isNested", "addNested",
        "modifierOptionPriceType", "isOptional", "canGuestSelectMoreModifiers",
        "multiSelect", "limitIndividualModifierSelection", "minSelector",
        "maxSelector", "noMaxSelection", "prefix", "pizzaSelection",
        "stockStatus", "price", "onPrem", "offPrem", "parentModifierId",
        "isSizeModifier",
    ],
    "Modifier Option": [
        "id", "optionName", "posDisplayName", "price", "isStockAvailable",
        "isSizeModifier",
    ],
    "Modifier ModifierOptions": [
        "modifierId", "modifierOptionId", "isDefaultSelected", "maxLimit",
        "optionDisplayName", "sortOrder",
    ],
    "Allergen": ["id", "allergenName", "iconId", "isDefault"],
    "Tag": ["id", "tagName", "iconId", "isDefault"],
}


def export_xlsx(model: dict, template_path: str, output_path: str):
    """
    Load template workbook, clear data rows, fill from model, save to output_path.
    """
    wb = openpyxl.load_workbook(template_path)

    for sheet_name, columns in SHEET_COLUMNS.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows = model.get(sheet_name, [])

        # Clear existing data rows (keep header row 1)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, col_key in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_key))

    wb.save(output_path)
